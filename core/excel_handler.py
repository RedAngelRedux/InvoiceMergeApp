import os
import re

import openpyxl

from core.email_builder import load_template, build_message, replace_placeholders
from core.mailer import send_email, archive_email
from core.utils.spinner import Spinner
from core.utils.uiux import match_file_in_folder, get_masked_input
from core.config_loader import IMAP_CONFIG

EMAIL_REGEX = re.compile(
    r"^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$"
)

def validate_emails(email_list):
    """
    Validates a list of email addresses.

    Args:
        email_list (list[str]): List of email strings.

    Returns:
        tuple: (True, []) if all valid, or (False, [invalid_emails])
    """
    invalid = [email for email in email_list if not EMAIL_REGEX.match(email)]
    return (len(invalid) == 0, invalid)

def find_attachment(folder, account):
    pattern = re.compile(rf"^{re.escape(account)}\sINV\s.*\.pdf$", re.IGNORECASE)
    for fname in os.listdir(folder):
        if pattern.match(fname):
            return os.path.join(folder, fname)
    return None

def load_sheet(file_name, tab_name):
    wb = openpyxl.load_workbook(file_name)
    if tab_name not in wb.sheetnames:
        raise ValueError(f"Tab '{tab_name}' not found.")
    return wb, wb[tab_name]

def read_email_rows(sheet, folder):
    from_email = sheet['A1'].value
    records = []
    for i, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=3):
        account, to, cc, bcc, status, archive = row
        if status and status.lower().startswith("sent"):
            continue
        attachment = find_attachment(folder, str(account))
        records.append({
            "account": account,
            "to": to.split(";") if to else [],
            "cc": cc.split(";") if cc else [],
            "bcc": bcc.split(";") if bcc else [],
            "status_row": i,
            "archive_row": i,
            "attachment": attachment
        })
    return from_email, records

def record_archive(sheet, row_index, result):
    col = "F"
    cell = sheet[f"{col}{row_index}"]
    cell.value = result
    if result.startswith("Error"):
        cell.font = openpyxl.styles.Font(color="FF0000")
    else:
        cell.font = openpyxl.styles.Font(color="000000")

def record_status(sheet, row_index, result):
    col = "E"
    cell = sheet[f"{col}{row_index}"]
    cell.value = result
    if result.startswith("Error"):
        cell.font = openpyxl.styles.Font(color="FF0000")
    else:
        cell.font = openpyxl.styles.Font(color="000000")

def email_all_invoices(EXCEL_FILE, tab, TIMESTAMPED_FOLDER):
    try:

        wb, sheet = load_sheet(EXCEL_FILE, tab)
        from_email, rows = read_email_rows(sheet, TIMESTAMPED_FOLDER)
        subject, body, signature = load_template(tab)

        # password = getpass.getpass(f"Enter the password for {from_email}:  ")
        password = get_masked_input(f"Enter the password for {from_email}:  ")

        with Spinner(f"Sending emails for group {tab}"):

            subject_template = subject
            body_template = body
            signature_template = signature
            for row in rows:        
                data = {
                    "ACCOUNT": row["account"]
                }
                subject = replace_placeholders(subject_template,data)
                body = replace_placeholders(body_template,data)
                signature = replace_placeholders(signature_template,data)

                still_good = True                
                error_parts = ["Error: "]
                attachment_exists = match_file_in_folder(TIMESTAMPED_FOLDER,str(row["account"]))
                if not attachment_exists:
                    error_parts.append("No Attachement Found; ")
                    still_good = False 
                    
                valid, bad_emails = validate_emails(row["to"])
                if not valid:
                    error_parts.append("Invalid E-mails in [TO] column - ")
                    error_parts.extend(bad_emails)
                    still_good = False

                valid, bad_emails = validate_emails(row["cc"])
                if not valid:
                    error_parts.append("Invalid E-mails in [CC] column - ")
                    error_parts.extend(bad_emails)
                    still_good = False

                valid, bad_emails = validate_emails(row["bcc"])
                if not valid:
                    error_parts.append("Invalid E-mails in [BCC] column - ")
                    error_parts.extend(bad_emails)
                    still_good = False

                if still_good:
                    msg = build_message(from_email, row["to"], row["cc"], row["bcc"], subject, body, signature, row["attachment"], password)
                    status, success = send_email(msg)
                    record_status(sheet,row["status_row"],status)
                    if success:
                        status, success = archive_email(msg,IMAP_CONFIG["archive_folder"])
                        record_archive(sheet,row["status_row"],status)                        
                    # comment, success = send_email(msg)
                    # if success:
                    #     arc_comment, arc_success = archive_email(msg)
                    #     if arc_success:
                    #         record_status(sheet, row["status_row"], comment.join(arc_comment))
                    #     else:
                    #         record_status(sheet, row["status_row"], comment.join(arc_comment))
                    # else:
                    #     record_status(sheet,row["status_row"],comment)
                else:
                    record_status(sheet,row["status_row"]," ".join(error_parts))

            wb.save(EXCEL_FILE)
    except Exception as e:
        raise
